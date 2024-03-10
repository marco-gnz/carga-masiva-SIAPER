from openpyxl import load_workbook
from lxml import etree
from datetime import datetime

def format_date(fecha_str):
    try:
        datetime_object = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
        return datetime_object.strftime('%d/%m/%Y')
    except ValueError:
        print(f"No se pudo convertir la cadena '{fecha_str}' al formato especificado.")
        return None


file_search_data    = 'carga.xlsx';
file_format_siaper  = 'formato_carga_masiva.xsd';
workbook            = load_workbook(filename=file_search_data)
sheet               = workbook.active


root = etree.Element("LISTADOCUMENTOS")

for row in sheet.iter_rows(min_row=2, values_only=True):
    documento = etree.SubElement(root, "DOCUMENTO")
    
    row = ["" if cell is None else cell for cell in row]
    etree.SubElement(documento, "RUN").text = str(row[0])
    etree.SubElement(documento, "DIGITO_VERIFICADOR").text = str(row[1])
    etree.SubElement(documento, "TIPO_DOCUMENTO").text = str(row[2])
    etree.SubElement(documento, "NUMERO_DOCUMENTO").text = str(row[3])
    
    fecha_excel_str = str(row[4])
    fecha_con_formato = format_date(fecha_excel_str)
    if fecha_con_formato:
        etree.SubElement(documento, "FECHA_DOCUMENTO").text = fecha_con_formato
        
    etree.SubElement(documento, "SERVICIO_EMISOR").text = str(row[5])
    etree.SubElement(documento, "DEPENDENCIA_EMISORA").text = str(row[6])
    etree.SubElement(documento, "SERVICIO_DESTINO").text = str(row[7])
    etree.SubElement(documento, "DEPENDENCIA_DESTINO").text = str(row[8])
    etree.SubElement(documento, "REGION_DESTINO").text = str(row[9])
    etree.SubElement(documento, "COMUNA_DESTINO").text = str(row[10])
    
    fecha_desde = str(row[11])
    fecha_hasta = str(row[12])
    fecha_desde_formato = format_date(fecha_desde)
    if fecha_desde_formato:
        etree.SubElement(documento, "FECHA_DESDE").text = fecha_desde_formato
        
    fecha_hasta_formato = format_date(fecha_hasta)
    if fecha_hasta_formato:
        etree.SubElement(documento, "FECHA_HASTA").text = fecha_hasta_formato
    
    etree.SubElement(documento, "MOTIVO_COMETIDO_FUNCIONARIO").text = str(row[13])

    tiene_beneficios = str(row[14])
    print("Valor de tiene_beneficios:", tiene_beneficios)
    
    elemento_tiene_beneficios = etree.SubElement(documento, "TIENE_BENEFICIOS")

    if tiene_beneficios != "No":
        elemento_seleccion_beneficios = etree.SubElement(elemento_tiene_beneficios, "SELECCIONE_BENEFICIOS")

        etree.SubElement(elemento_seleccion_beneficios, "PASAJE").text = str(row[14])

        etree.SubElement(elemento_seleccion_beneficios, "VIATICO").text = str(row[15])
        
        etree.SubElement(elemento_seleccion_beneficios, "ALOJAMIENTO").text = str(row[16])
        
        otros_value = str(row[17])
        if otros_value:
            elemento_otros = etree.SubElement(elemento_seleccion_beneficios, "OTROS")
            etree.SubElement(elemento_otros, "DETALLE_BENEFICIOS_PERIODO").text = otros_value

        monto_value = row[19]
        if monto_value is not None:
            etree.SubElement(documento, "MONTO").text = str(monto_value)

    else:
        #Si no hay beneficios, simplemente establecer el texto de TIENE_BENEFICIOS
        elemento_tiene_beneficios.text = "No"
        
    xml_tree = etree.ElementTree(root)
    xml_schema = etree.XMLSchema(file=file_format_siaper)

if xml_schema.validate(xml_tree):
    xml_tree.write('cometidos.xml', pretty_print=True)
    print("El documento XML ha sido generado correctamente.")
else:
    print("El documento XML no cumple con el esquema XSD.")
    print(xml_schema.error_log)
